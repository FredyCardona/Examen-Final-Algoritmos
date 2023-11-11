import openpyxl

def crear_vehiculo():
    # Solicitar al usuario ingresar los datos
    codigo = input("Ingrese el código del vehículo: ")
    marca = input("Ingrese la marca del vehículo: ")
    modelo = input("Ingrese el modelo del vehículo: ")
    precio = float(input("Ingrese el precio del vehículo: "))
    kilometraje = int(input("Ingrese el kilometraje del vehículo: "))

    # Crear un nuevo vehículo
    vehiculo = {
        "Código": codigo,
        "Marca": marca,
        "Modelo": modelo,
        "Precio": precio,
        "Kilometraje": kilometraje
    }
    return vehiculo

def guardar_vehiculos(vehiculos):
    # Guardar la lista de vehículos en el archivo Excel
    libro = openpyxl.load_workbook('vehiculos.xlsx')
    hoja = libro['listado']

    for vehiculo in vehiculos:
        hoja.append(list(vehiculo.values()))

    libro.save('vehiculos.xlsx')

def listar_vehiculos():
    # Listar todos los vehículos en el archivo Excel
    libro = openpyxl.load_workbook('vehiculos.xlsx')
    hoja = libro['listado']

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        print(dict(zip(["Código", "Marca", "Modelo", "Precio", "Kilometraje"], fila)))

def eliminar_vehiculo():
    # Solicitar al usuario el código del vehículo a eliminar
    codigo = input("Ingrese el código del vehículo a eliminar: ")

    # Eliminar el vehículo del archivo Excel
    libro = openpyxl.load_workbook('vehiculos.xlsx')
    hoja = libro['listado']

    for fila in hoja.iter_rows(min_row=2, max_col=1, values_only=True):
        if fila[0] == codigo:
            hoja.delete_rows(fila[0].row)
            break

    libro.save('vehiculos.xlsx')

# Puedes agregar más funciones según sea necesario (editar, cargar masivamente, etc.)

# Ejemplo de uso
vehiculos = []

# Solicitar al usuario crear varios vehículos
num_vehiculos = int(input("Ingrese el número de vehículos a ingresar: "))
for _ in range(num_vehiculos):
    vehiculo = crear_vehiculo()
    vehiculos.append(vehiculo)

# Guardar y listar los vehículos
guardar_vehiculos(vehiculos)
listar_vehiculos()

# Solicitar al usuario eliminar un vehículo
eliminar_vehiculo()
listar_vehiculos()
